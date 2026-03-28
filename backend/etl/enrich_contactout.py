# etl/enrich_contactout.py
#
# ContactOut email enrichment — CSV export / Excel import workflow.
#
# ContactOut's Data Enrichment does not expose a public API.
# It uses a manual upload/download workflow:
#
#   Step 1 — EXPORT (--export mode)
#   --------------------------------
#   Generates exports/contactout_upload.csv formatted for ContactOut's
#   "Get Work Emails" template: First Name, Last Name, Domain.
#   Only physicians with a practice_domain are included.
#   Ordered by lead_score_current DESC — best leads consume credits first.
#
#   Step 2 — MANUAL UPLOAD (you do this in the browser)
#   -----------------------------------------------------
#   1. Go to: contactout.com/dashboard/data-enrichment
#   2. Click: "Get Work Emails"
#   3. Upload: exports/contactout_upload.csv
#   4. Wait for enrichment to complete
#   5. Download the results file (ContactOut returns .xlsx)
#   6. Save it as: exports/contactout_results.xlsx
#
#   Step 3 — IMPORT (--import mode)
#   ---------------------------------
#   Reads exports/contactout_results.xlsx, saves verified emails
#   to personal_email, updates enrichment_sources[],
#   updates enrichment_source_stats, syncs leads table.
#
# ContactOut results column layout (confirmed from live test)
# -----------------------------------------------------------
#   Column 1: First name
#   Column 2: Last name
#   Column 3: Domain
#   Column 4: Work email       — contains NPI number for ALL rows (their bug/quirk)
#   Column 5: Work email status — contains the email address on HIT rows,
#                                 'No match found.' on MISS rows
#   Column 6: (unnamed)        — contains 'Verified' on HIT rows, None on MISS rows
#
#   Detection logic:
#     - Scan ALL columns in each row for a valid email address (@domain)
#     - If found anywhere → it's a hit row, use that email
#     - If not found → no match
#
# Usage
# -----
#   python etl/enrich_contactout.py --export           # generate upload CSV
#   python etl/enrich_contactout.py --export --limit 5 # top N only (default 5)
#   python etl/enrich_contactout.py --import           # process results Excel
#
# Credit limit
# ------------
#   ContactOut free tier: 5 email credits per day (daily reset).
#   Always run --export --limit 5 to stay within free tier.
#
# Dependencies
# ------------
#   pip install openpyxl

import sys
import csv
import argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).parent.parent))

from database import engine

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

SOURCE_NAME  = "contactout"
EXPORTS_DIR  = Path(__file__).parent.parent / "exports"
UPLOAD_FILE  = EXPORTS_DIR / "contactout_upload.csv"
RESULTS_FILE = EXPORTS_DIR / "contactout_results.xlsx"

# Known no-match status strings from ContactOut
NO_MATCH_STRINGS = {"no match found.", "no match found", "not found"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ensure_exports_dir() -> None:
    """Creates the exports/ directory if it does not exist."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def is_valid_email(value) -> bool:
    """
    Returns True if value looks like an email address.

    Handles string and non-string values safely.
    ContactOut puts NPI numbers (integers) in the Work email column —
    this check filters those out correctly.

    Args:
        value: Any value from a spreadsheet cell.

    Returns:
        True if value is a string containing '@' with a valid domain part.
    """
    if not isinstance(value, str):
        return False
    value = value.strip()
    return "@" in value and "." in value.split("@")[-1]


def extract_email_from_row(row_values: tuple) -> Optional[str]:
    """
    Scans all cells in a row for a valid email address.

    ContactOut's column layout is inconsistent — the email appears
    in different columns depending on hit/miss status. Rather than
    hardcoding a column index, we scan all cells and take the first
    valid email found anywhere in the row.

    Args:
        row_values: Tuple of cell values from openpyxl iter_rows.

    Returns:
        Email string if found, None otherwise.
    """
    for cell_value in row_values:
        if is_valid_email(cell_value):
            return str(cell_value).strip()
    return None


# ---------------------------------------------------------------------------
# Export mode
# ---------------------------------------------------------------------------

def run_export(limit: int = 5) -> None:
    """
    Generates a ContactOut-formatted CSV for manual upload.

    Pulls physicians that:
        - Have no personal_email or legacy email yet
        - Have email_enrichment_attempted = FALSE
        - Have a practice_domain (required by ContactOut)
        - Are active

    Ordered by lead_score_current DESC — best leads consume credits first.

    Output: exports/contactout_upload.csv
    Columns: First Name, Last Name, Domain, npi

    The 'npi' column is extra — ContactOut ignores unknown columns
    but we keep it as a reference for debugging the import step.

    Args:
        limit: Max rows to export. Default 5 = ContactOut free daily limit.
    """
    ensure_exports_dir()

    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                p.npi,
                p.first_name_clean,
                p.last_name_clean,
                p.practice_domain,
                p.lead_score_current
            FROM physician p
            WHERE p.personal_email IS NULL
              AND p.email IS NULL
              AND p.is_active = TRUE
              AND NOT (enrichment_sources @> ARRAY['contactout']::text[])
              AND p.practice_domain IS NOT NULL
            ORDER BY p.lead_score_current DESC NULLS LAST
            LIMIT :limit
        """), {"limit": limit})
        rows = result.fetchall()

    if not rows:
        print("No eligible physicians found for ContactOut export.")
        print("Possible reasons:")
        print("  - All physicians with domains already attempted")
        print("  - No physicians have practice_domain populated")
        return

    with open(UPLOAD_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["First Name", "Last Name", "Domain", "npi"])
        for row in rows:
            writer.writerow([
                row[1] or "",  # first_name_clean
                row[2] or "",  # last_name_clean
                row[3] or "",  # practice_domain
                row[0],        # npi — ContactOut ignores, kept for reference
            ])

    print("=" * 60)
    print("CONTACTOUT EXPORT COMPLETE")
    print(f"  File    : {UPLOAD_FILE}")
    print(f"  Rows    : {len(rows)}")
    print()
    print("NEXT STEPS:")
    print("  1. Go to : contactout.com/dashboard/data-enrichment")
    print("  2. Click : 'Get Work Emails'")
    print("  3. Upload:", str(UPLOAD_FILE))
    print("  4. Wait for enrichment to finish")
    print("  5. Download the results Excel file")
    print(f"  6. Save as: {RESULTS_FILE}")
    print("  7. Run   : python etl/enrich_contactout.py --import")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Import mode
# ---------------------------------------------------------------------------

def load_results_excel() -> Optional[list[tuple]]:
    """
    Loads the ContactOut results Excel file as a list of raw row tuples.

    Returns raw tuples (not dicts) so extract_email_from_row can scan
    all cells regardless of column position.

    Returns:
        List of (first_name, last_name, domain, *other_cells) tuples,
        or None if file missing or openpyxl not installed.
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl not installed.")
        print("Run: pip install openpyxl --break-system-packages")
        return None

    if not RESULTS_FILE.exists():
        print(f"ERROR: Results file not found: {RESULTS_FILE}")
        print("Complete the manual upload steps first, then re-run --import.")
        return None

    wb   = openpyxl.load_workbook(RESULTS_FILE)
    ws   = wb.active
    if ws is None:
        print("Error: worksheet is empty or cannot be read.")
        return None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header row
        rows.append(row)
    return rows


def resolve_npi(first: str, last: str, domain: str, conn) -> Optional[str]:
    """
    Resolves physician NPI from name and domain.

    Matching strategy:
        1. Domain match — most reliable (domain is unique per physician)
        2. Name match  — fallback if domain not in DB

    Args:
        first:  First name from ContactOut row.
        last:   Last name from ContactOut row.
        domain: Practice domain from ContactOut row.
        conn:   Active SQLAlchemy connection.

    Returns:
        NPI string if found, None otherwise.
    """
    if domain:
        result = conn.execute(text("""
            SELECT npi FROM physician
            WHERE practice_domain = :domain
            LIMIT 1
        """), {"domain": domain})
        match = result.fetchone()
        if match:
            return match[0]

    if first and last:
        result = conn.execute(text("""
            SELECT npi FROM physician
            WHERE LOWER(first_name_clean) = LOWER(:first)
              AND LOWER(last_name_clean)  = LOWER(:last)
            LIMIT 1
        """), {"first": first, "last": last})
        match = result.fetchone()
        if match:
            return match[0]

    return None


def run_import() -> None:
    """
    Processes the ContactOut results Excel file and saves emails to DB.

    Detection strategy:
        Scans every cell in each row for a valid email address.
        ContactOut's column layout is inconsistent — the email can
        appear in different columns. Scanning all cells is robust
        against their quirky output format.

    For each row where a valid email is found:
        1. Saves to physician.personal_email (canonical)
        2. Sets personal_email_confidence = 'HIGH'
        3. Appends 'contactout' to physician.enrichment_sources[]
        4. Sets email_enrichment_attempted = TRUE
        5. Inserts audit row in field_value_history
        6. Updates enrichment_source_stats atomically

    For each no-match row:
        1. Sets email_enrichment_attempted = TRUE
        2. Increments total_attempts in stats

    After all rows: syncs leads table.
    """
    now = datetime.now(timezone.utc)

    print("=" * 60)
    print("CONTACTOUT IMPORT")
    print(f"  File: {RESULTS_FILE}")
    print("=" * 60)

    rows = load_results_excel()
    if rows is None:
        return
    if not rows:
        print("Results file is empty — nothing to import.")
        return

    print(f"  Rows in file: {len(rows)}")

    saved     = 0
    not_found = 0
    no_match  = 0
    errors    = 0

    with engine.connect() as conn:
        for i, row_values in enumerate(rows, 1):

            # ContactOut columns: First name(0), Last name(1), Domain(2), ...
            first  = str(row_values[0] or "").strip()
            last   = str(row_values[1] or "").strip()
            domain = str(row_values[2] or "").strip()

            # Scan ALL cells for a valid email
            email = extract_email_from_row(row_values)

            print(f"\n  [{i}] {first} {last} | {domain}")
            print(f"       Raw row: {row_values}")

            # ── Resolve NPI ────────────────────────────────────────────
            npi = resolve_npi(first, last, domain, conn)
            if not npi:
                print(f"       Cannot match to physician — skipping")
                no_match += 1
                continue

            print(f"       NPI: {npi}")

            # ── No email found ─────────────────────────────────────────
            if not email:
                print(f"       No email found in any column")
                not_found += 1
                conn.execute(text("""
                    UPDATE physician SET
                        enrichment_sources = CASE
                            WHEN enrichment_sources @> ARRAY['contactout']::text[]
                            THEN enrichment_sources
                            ELSE array_append(enrichment_sources, 'contactout')
                        END,
                        enrichment_last_attempted_at = :now,
                        updated_at                   = :now
                    WHERE npi = :npi
                """), {"npi": npi, "now": now})
                conn.execute(text("""
                    UPDATE enrichment_source_stats SET
                        total_attempts = total_attempts + 1,
                        last_used_at   = :now,
                        updated_at     = :now
                    WHERE source_name  = :source
                """), {"source": SOURCE_NAME, "now": now})
                conn.commit()
                continue

            # ── Valid email found ──────────────────────────────────────
            print(f"       Email found: {email}")

            try:
                conn.execute(text("""
                    UPDATE physician SET
                        personal_email             = :email,
                        personal_email_confidence  = 'HIGH',
                        email_enriched_at          = :now,
                        email_enrichment_attempted = TRUE,
                        email_enrichment_result    = 'contactout_found',
                        enrichment_sources = CASE
                            WHEN enrichment_sources @> ARRAY[:source]::text[]
                            THEN enrichment_sources
                            ELSE array_append(enrichment_sources, :source)
                        END,
                        enrichment_last_attempted_at = :now,
                        updated_at                 = :now
                    WHERE npi = :npi
                """), {
                    "npi":    npi,
                    "email":  email,
                    "source": SOURCE_NAME,
                    "now":    now,
                })

                conn.execute(text("""
                    INSERT INTO field_value_history (
                        history_id, entity_type, entity_id, npi,
                        field_name, field_value, source_name,
                        confidence_score, is_current,
                        collected_timestamp, created_at
                    ) VALUES (
                        gen_random_uuid(), 'physician', :npi, :npi,
                        'personal_email', :email, :source,
                        0.90, TRUE, :now, :now
                    )
                """), {
                    "npi": npi, "email": email,
                    "source": SOURCE_NAME, "now": now,
                })

                conn.execute(text("""
                    UPDATE enrichment_source_stats SET
                        emails_provided = emails_provided + 1,
                        total_hits      = total_hits      + 1,
                        total_attempts  = total_attempts  + 1,
                        last_used_at    = :now,
                        updated_at      = :now
                    WHERE source_name   = :source
                """), {"source": SOURCE_NAME, "now": now})

                conn.commit()
                saved += 1
                print(f"       SAVED.")

            except Exception as e:
                print(f"       ERROR: {e}")
                errors += 1
                conn.rollback()

    _sync_leads(now)

    total = len(rows)
    print()
    print("=" * 60)
    print("CONTACTOUT IMPORT COMPLETE")
    print(f"  Rows processed : {total}")
    print(f"  Emails saved   : {saved}")
    print(f"  No email found : {not_found}")
    print(f"  No NPI match   : {no_match}")
    print(f"  Errors         : {errors}")
    if total > 0:
        print(f"  Hit rate       : {saved / total * 100:.1f}%")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Leads sync
# ---------------------------------------------------------------------------

def _sync_leads(now: datetime) -> None:
    """
    Syncs physicians with contact data into the leads table.

    contact_category:
        A = mobile_phone + any email
        B = any email, no mobile_phone
        Rows with no contact info never inserted.

    Idempotent — safe to call multiple times.
    """
    print("\nSyncing leads table...")
    with engine.connect() as conn:
        conn.execute(text("""
            INSERT INTO leads (
                npi, first_name, last_name, full_name, credential,
                specialty, specialty_category, organization_name,
                practice_domain,
                email, personal_email, personal_email_confidence,
                practice_email,
                email_confidence_score, email_confidence_level,
                email_verification_status, email_source,
                address_line_1, city, state, zip,
                lead_score, lead_tier,
                years_of_experience, experience_bucket,
                license_count, multi_state_flag,
                mobile_phone, phone_confidence,
                contact_completeness,
                contact_category,
                enrichment_sources,
                created_at, updated_at
            )
            SELECT
                p.npi, p.first_name_clean, p.last_name_clean,
                p.full_name_display, p.credential_normalized,
                p.specialty_name, p.derived_specialty_category,
                p.organization_name, p.practice_domain,
                p.email, p.personal_email, p.personal_email_confidence,
                p.practice_email,
                p.email_confidence_score, p.email_confidence_level,
                p.email_verification_status, p.email_source,
                ppl.address_line_1, ppl.city, ppl.state, ppl.zip,
                p.lead_score_current, p.lead_tier,
                p.years_of_experience, p.experience_bucket,
                p.license_count, p.multi_state_flag,
                p.mobile_phone, p.phone_confidence,
                p.contact_completeness,
                CASE
                    WHEN p.mobile_phone IS NOT NULL
                     AND (
                           p.personal_email IS NOT NULL
                        OR p.practice_email IS NOT NULL
                        OR p.email          IS NOT NULL
                     )
                    THEN 'A'
                    WHEN (
                           p.personal_email IS NOT NULL
                        OR p.practice_email IS NOT NULL
                        OR p.email          IS NOT NULL
                    )
                    THEN 'B'
                    ELSE NULL
                END,
                p.enrichment_sources,
                :now, :now
            FROM physician p
            LEFT JOIN physician_practice_locations ppl
                ON p.npi = ppl.npi
               AND ppl.is_primary_location = TRUE
            WHERE (
                p.personal_email IS NOT NULL
                OR p.practice_email IS NOT NULL
                OR p.email IS NOT NULL
            )
            ON CONFLICT (npi) DO UPDATE SET
                personal_email            = EXCLUDED.personal_email,
                personal_email_confidence = EXCLUDED.personal_email_confidence,
                email                     = EXCLUDED.email,
                email_confidence_score    = EXCLUDED.email_confidence_score,
                email_confidence_level    = EXCLUDED.email_confidence_level,
                email_verification_status = EXCLUDED.email_verification_status,
                mobile_phone              = EXCLUDED.mobile_phone,
                phone_confidence          = EXCLUDED.phone_confidence,
                lead_score                = EXCLUDED.lead_score,
                lead_tier                 = EXCLUDED.lead_tier,
                contact_category          = EXCLUDED.contact_category,
                enrichment_sources        = EXCLUDED.enrichment_sources,
                updated_at                = EXCLUDED.updated_at
        """), {"now": now})
        conn.commit()

        result = conn.execute(text("SELECT COUNT(*) FROM leads"))
        count_row = result.fetchone()
        count  = int(count_row[0]) if count_row and count_row[0] is not None else 0
        print(f"  Total leads: {count}")

        cats = conn.execute(text("""
            SELECT contact_category, COUNT(*) AS cnt
            FROM leads GROUP BY contact_category ORDER BY contact_category
        """))
        for r in cats.fetchall():
            print(f"  Category {r[0]}: {r[1]}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="ContactOut Email Enrichment — CSV Export / Excel Import"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--export", action="store_true",
        help="Generate CSV for ContactOut upload"
    )
    group.add_argument(
        "--import", dest="do_import", action="store_true",
        help="Process ContactOut results Excel file"
    )
    parser.add_argument(
        "--limit", type=int, default=5,
        help="Max physicians to export (default: 5 = ContactOut free daily limit)"
    )
    args = parser.parse_args()

    if args.export:
        run_export(limit=args.limit)
    elif args.do_import:
        run_import()